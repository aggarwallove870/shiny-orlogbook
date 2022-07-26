
from cmath import e
from genericpath import exists
from apps.authentication.models import *
from apps.home.models import *
from typing import Counter
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import redirect, render
from django.template import loader
from .models import Graphs
import pandas as pd
from django.contrib import messages
from datetime import timedelta,date
import os
from django.http import FileResponse
from openpyxl import load_workbook
from django.contrib.auth.hashers import make_password


@login_required(login_url="/login/")
def index(request):
    context = {'segment': 'index'}

    html_template = loader.get_template('home/index.html')
    return HttpResponse(html_template.render(context, request))


def read_data(file_path):
    
            pandas_data = pd.read_excel(file_path, sheet_name='RAW', header=2, parse_dates=True)
            pandas_data_frame= pandas_data.sort_values(by='Date')
            data_frame = pd.DataFrame(pandas_data_frame)
            return data_frame   
    
        
       


def color_line_chart(key):
    if key == "Primary Surgeon":
        color = '#398AB9'
    elif key == "First Assist":
        color = '#D8D2CB'
    elif key == "Secondary Assist":
        color = '#1A374D'

    return color


@login_required(login_url="/login/")
def generate_bar_chart(request):


      
    if request.method == 'GET':

        institute = Institution.objects.all()
        supervisor = Supervisor.objects.all()
        if not request.user.is_superuser and not request.user.is_supervisor:
            obj = Graphs.objects.filter(user=request.user)
        elif request.method == "GET" and request.user.is_superuser or request.user.is_supervisor:
            if request.GET.get('user_id'):
                obj = Graphs.objects.filter(user=int(request.GET.get('user_id')))
            else:
                obj = Graphs.objects.filter(user=int(77777))
           
        else:
            obj = Graphs.objects.filter(user=int(7777))

        if obj.exists():
            obj = obj.last()
            institute = Institution.objects.all()
            supervisor = Supervisor.objects.all()
        
            file_path = obj.upload.path
            data = read_data(file_path)


            context_di = {}
            staff = {}
            
            for i in data['Staff']:
                context_di[i] =  data[data['Staff'] == i]


                staff[i] = dict(Counter(context_di[i]['Role']))

    
            a = data['Staff']
            get_staff = list(a.unique())

            b = data['Sub-Specialty']
            get_sub_specialty = list(b.unique())

            c = data['Location']
            get_location = list(c.unique())

            d = data['Role']
            get_role_data = list(d.unique())

            e = data['PGY']
            get_pgy = list(e.unique())
    

            total_cases = len(data)
            current_year = date.today().year
            get_date = data[data['Date'].dt.year == current_year] 
            date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
            count_of_current_year = 0
            for u in list(date_gt):
                if u == True:
                    count_of_current_year+=1
                else:
                    pass
           
            date_equal =  data[data['Date'] == str(current_year)+'-'+'07-01'] 
            this_year_ps = date_equal['Role'] == 'Primary Surgeon' 
            count_of_current = 0
            for i in this_year_ps:
                if i == True:
                    count_of_current+=1
                else:
                    pass
            today = date.today()
            first = today.replace(day=1)
            lastMonth = first - timedelta(days=1)
            last_months = lastMonth.strftime("%Y-%m")
            this_year_month = data[data['Date'].dt.to_period('M') == last_months]
            get_g = this_year_month['Role'] == 'Primary Surgeon'
            count_of_last_month = 0
            for i in get_g:
                if i == True:
                    count_of_last_month+=1
                else:
                    pass


            today = date.today()
            datem = today.strftime("%Y-%m")
        
            this_year_this_month = data[data['Date'].dt.to_period('M') == datem]
            get_role = this_year_this_month['Role'] == 'Primary Surgeon'
            count_of_this_month = 0
            for i in get_role:
                if i == True:
                    count_of_this_month+=1
                else:
                    pass

            coded_case = []
            for i in data['Coded Case']:

                coded_case.append(i)

            
            dashboard113 = (dict(Counter(coded_case)))
            dashboard23 = (dict(sorted(dashboard113.items(), key=lambda x:x[1], reverse=True)))
            
          
         
            datasets = [
                {
                "label": 'Primary Surgeon',
                "backgroundColor": '#398AB9',
                "data": []
                },
                {
                "label": 'First Assist',
                "backgroundColor": '#D8D2CB',
                "data": []
                },
                {
                "label": 'Secondary Assist',
                "backgroundColor": '#1A374D',
                "data": []
                }
            ]


            only_final_data = []

            for d in staff:
                final_data = {
                    "name": d,
                    "Primary Surgeon": staff[d].get('Primary Surgeon', 0),
                    "First Assist": staff[d].get('First Assist', 0),
                    "Secondary Assist": staff[d].get('Secondary Assist', 0),
                }
            
                only_final_data.append(final_data)
          
        
            lists = []
            
            for i in only_final_data:
                total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                i['Total'] = total
                lists.append(i)
            
   
            shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)
            name_set = []
            for i in shorted_data:
                name_set.append(i['name'])

            p_list = []
            f_list = []
            s_list = []
            for d in shorted_data:
                p_list.append(d['Primary Surgeon']) 
                f_list.append(d['First Assist']) 
                s_list.append(d['Secondary Assist']) 
                # s_list.append(d['Secondary Assist']) 

            
            for dataset in datasets:
                if dataset['label'] == 'Primary Surgeon':
                    dataset['data'] = p_list
                elif dataset['label'] == 'First Assist':
                    dataset['data'] = f_list
                elif dataset['label'] == 'Secondary Assist':
                    dataset['data'] = s_list

            
            role = []
            for i in data['Role']:
                role.append(i)
            
            dashboard1 = dict(Counter(role))
            keys = list(dashboard1.keys())
            values = list(dashboard1.values())

    #add new code here
            specialty_chart = {}
            roles = {}
            for i in data['Role']:
                specialty_chart[i] = data[data['Role'] == i]
                roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

            get_data = []
          
            for key, value in roles.items():
        
                dat = {
                    "label": key,
                    "Trauma": value.get("Trauma",0),
                    "Arthroplasty": value.get("Arthroplasty",0),
                    "Pediatrics": value.get("Pediatrics",0),
                    "Sports": value.get("Sports",0),
                    "Foot and Ankle": value.get("Foot and Ankle",0),
                    "Upper Extremity": value.get("Upper Extremity",0),
                    "General": value.get("General",0),
                    "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                }
                get_data.append(dat)

            shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)
          
            
            key1 = list(roles.values())
            labels_data = list(key1[0].keys())
        
         
            grt_Data = []

            for i in shorted_data_list:
                d ={
                        "label": i.get("label"),
                        "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                        "backgroundColor": color_line_chart(i.get("label"))
                    }
                grt_Data.append(d)    


            # Driver code by kapil START
            raw_list = []
            polished_data = []
            total_roles = [] # Here storing roles
            total_sub = [] # Here storing all the speiclaity
            for single_role in roles.keys():
                total_roles.append(single_role)
    
            for sub in total_roles:
                for m in roles[sub].keys():
                    if m not in total_sub:
                        total_sub.append(m)

            i = 0
            while i<len(total_sub):
                final_data = {
                                "name": total_sub[i],
                                # "Primary Surgeon": roles['Primary Surgeon'].get(total_sub[i],0),
                                "Primary Surgeon": roles.get('Primary Surgeon', {}).get(total_sub[i],0),
                                "First Assist": roles.get('First Assist',{}).get(total_sub[i],0),
                                "Secondary Assist": roles['Secondary Assist'].get(total_sub[i],0)
                            }
                raw_list.append(final_data)
                i=i+1
                

            for i in raw_list:
                
                total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                i['Total'] = total
                polished_data.append(i)

            shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
            p_spec_list = []
            f_spec_list = []
            s_spec_list = []
            for d in shorted_data_list:
                p_spec_list.append(d['Primary Surgeon']) 
                f_spec_list.append(d['First Assist']) 
                s_spec_list.append(d['Secondary Assist']) 

            datas = [
                {
                "label": 'Primary Surgeon',
                "backgroundColor": '#398AB9',
                "data": []
                },
                {
                "label": 'First Assist',
                "backgroundColor": '#D8D2CB',
                "data": []
                },
                {
                "label": 'Secondary Assist',
                "backgroundColor": '#1A374D',
                "data": []
                }
            ]
            for dataset in datas:
                if dataset['label'] == 'Primary Surgeon':
                    dataset['data'] = p_spec_list
                elif dataset['label'] == 'First Assist':
                    dataset['data'] = f_spec_list
                elif dataset['label'] == 'Secondary Assist':
                    dataset['data'] = s_spec_list

            # Driver code by kapil END

            #site bar chart       
            
            site_chart = {}
            roles_list = {}
            for i in data['Role']:
                site_chart[i] = data[data['Role'] == i]
                roles_list[i] = dict(Counter(site_chart[i]['Location']))
            

            get_data = []
          
            for key, value in roles_list.items():
        
                dat = {
                    "label": key,
                    "HGH": value.get("HGH",0),
                    "JH": value.get("JH",0),
                    "MUMC": value.get("MUMC",0),
                    "SJH": value.get("SJH",0),
                }
                get_data.append(dat)
            
        

            key1 = list(roles_list.values())
            labels_site = list(key1[0].keys())
 
            
         
            granted_Data = []

            for i in get_data:
                d ={
                        "label": i.get("label"),
                        "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                        "backgroundColor": color_line_chart(i.get("label"))
                    }
                granted_Data.append(d) 


            ## Driver code by kapil START
            raw_site_data = []
            polish_site_data = []
            roles_site = []
            total_site = []
            for role_sit in roles_list.keys():
                roles_site.append(role_sit)
    
            for sub in roles_site:
                for m in roles_list[sub].keys():
                    if m not in total_site:
                        total_site.append(m)

            i = 0
            while i<len(total_site):
                final_site_data = {
                    "name": total_site[i],
                    # "Primary Surgeon": roles_list['Primary Surgeon'].get(total_site[i],0),
                    "Primary Surgeon": roles_list.get('Primary Surgeon', {}).get(total_sub[i],0),
                    "First Assist": roles_list.get('First Assist',{}).get(total_site[i],0),
                    "Secondary Assist": roles_list['Secondary Assist'].get(total_site[i],0)
                }
                raw_site_data.append(final_site_data)
                i= i+1

            for i in raw_site_data:
                total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                i['Total'] = total
                polish_site_data.append(i)

            shorted_site_data = sorted(polish_site_data, key=lambda x: x['Total'], reverse=True)
            p_site_list = []
            f_site_list = []
            s_site_list = []
            for d in shorted_site_data:
                p_site_list.append(d['Primary Surgeon']) 
                f_site_list.append(d['First Assist']) 
                s_site_list.append(d['Secondary Assist']) 

            datas_site = [
                {
                "label": 'Primary Surgeon',
                "backgroundColor": '#398AB9',
                "data": []
                },
                {
                "label": 'First Assist',
                "backgroundColor": '#D8D2CB',
                "data": []
                },
                {
                "label": 'Secondary Assist',
                "backgroundColor": '#1A374D',
                "data": []
                }
            ]
            for dataset in datas_site:
                if dataset['label'] == 'Primary Surgeon':
                    dataset['data'] = p_site_list
                elif dataset['label'] == 'First Assist':
                    dataset['data'] = f_site_list
                elif dataset['label'] == 'Secondary Assist':
                    dataset['data'] = s_site_list

            ## Driver code by kapil END
            
   

   #######
            
     
            context_dict = {}
            roles = {}
            # print(data['Date'].dt.strftime('%b-%Y'))
            for i in data['Date'].dt.strftime('%b-%Y'):
                context_dict[i] =  data[data['Date'].dt.strftime('%b-%Y') == i]
                roles[i] = dict(Counter(context_dict[i]['Role']))
            
        
            _data = {}
               
            for r in roles:
                _data_set = {
                    'Secondary Assist': roles[r].get('Secondary Assist', 0),
                    'First Assist': roles[r].get('First Assist', 0),
                    'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                }
                
                _data[r] = _data_set
        
            labels = list(roles.keys())
            role_keys = []
            raw_data = []
            for role in _data:
                role_keys = (list(_data[role].keys()))
                raw_data.append(_data[role])
            final_data = []
            for key in role_keys:
                data = []
                for raw in raw_data:
                    data.append(raw[key])
                    
                final_data.append({
                "label": key,
                "data": data,
                "borderColor": color_line_chart(key),
                "fill": "true",
                })
        
            final_final_data = {
                "labels": labels,
                
                "datasets": final_data,   
            }

            try:
                if request.user.is_superuser:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                else:
                    sups = Supervisor.objects.filter(username=request.user.username)[0]
                    users = NewUser.objects.filter(supervisor=sups.id)
            except:
                users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
          

            context = { 'keys':keys , 'values':values,'sub_speciality_label':total_sub,
            "labels_data":total_sub,
            "final_data_specialty_chart":datas,
            "granted_Data":datas_site,
            "labels_site":total_site,
             'final_data': final_final_data,
              'labels': labels, 'dashboard23':dashboard23,
               "total_cases":total_cases, "count_of_current_year":count_of_current_year,
                "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 
                'count_of_this_month':count_of_this_month,"final_data_list":datasets,"names":name_set,
                 "get_staff":get_staff,"get_role_data":get_role_data,
                  "get_sub_specialty":get_sub_specialty, "get_location":get_location, 
                  "users":users,"get_pgy":get_pgy,
                  'institute':institute, 'supervisor':supervisor}
          

            return render(request, 'home/sample.html', context)
        else:
            try:
                if request.user.is_superuser:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                else:
                    sups = Supervisor.objects.filter(username=request.user.username)[0]
                    users = NewUser.objects.filter(supervisor=sups.id)
            except:
                users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)


            return render(request, 'home/sample.html',{'users':users, 'institute':institute, 'supervisor':supervisor})
    
    if request.method == 'POST':

        institute = Institution.objects.all()
        supervisor = Supervisor.objects.all()
        if not request.FILES.get('document') and not request.user.is_superuser and  not request.user.is_supervisor:
            obj = Graphs.objects.filter(user=request.user)
            if obj.exists():
                obj = obj.last()
            
                upload_file = obj.upload.path
        

        elif not request.FILES.get('document') and request.user.is_superuser and  not request.user.is_supervisor:
            obj = Graphs.objects.filter(user=request.GET.get('user_id'))
            if obj.exists():
                obj = obj.last()
            
                upload_file = obj.upload.path     

        elif not request.FILES.get('document') and request.user.is_supervisor and not request.user.is_superuser :

            obj = Graphs.objects.filter(user=request.GET.get('user_id'))
            if obj.exists():
                obj = obj.last()
            
                upload_file = obj.upload.path     
        else:
            upload_file = request.FILES['document']
            try:
                wb = load_workbook(upload_file, read_only=True) 
                if 'RAW' not in wb.sheetnames:  
                    messages.error(request, "Please upload a valid file")
                    return render(request,'home/sample.html')
            except:
                pass
            file_extension = os.path.splitext(upload_file.name)[1]
            valid_extensions = [".xlsx", ".XLSX", ".xls", ".XLS"]

            if not file_extension.lower() in valid_extensions:
                msg = "Invalid file, select a valid xlsx file"
                messages.error(request, msg)
                return render(request,'home/sample.html')   

        data = read_data(upload_file)

        total_cases = len(data)
        current_year = date.today().year
        get_date = data[data['Date'].dt.year == current_year] 

        date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
        count_of_current_year = 0
        for u in list(date_gt):
            if u == True:
                count_of_current_year+=1
            else:
                pass
        date_equal =  data[data['Date'] == str(current_year)+'-'+'07-01'] 
        this_year_ps = date_equal['Role'] == 'Primary Surgeon' 

        count_of_current = 0
        for i in this_year_ps:
            if i == True:
                count_of_current+=1
            else:
                pass
        today = date.today()
        first = today.replace(day=1)
        lastMonth = first - timedelta(days=1)
        last_months = lastMonth.strftime("%Y-%m")

        this_year_month = data[data['Date'].dt.to_period('M') == last_months]
        get_g = this_year_month['Role'] == 'Primary Surgeon'
        count_of_last_month = 0
        for i in get_g:
            if i == True:
                count_of_last_month+=1
            else:
                pass


        today = date.today()
        datem = today.strftime("%Y-%m")
        this_year_this_month = data[data['Date'].dt.to_period('M') == datem]
        get_role = this_year_this_month['Role'] == 'Primary Surgeon'
        count_of_this_month = 0
        for i in get_role:
            if i == True:
                count_of_this_month+=1
            else:
                pass

        a = data['Staff']

        get_staff = list(a.unique())

        b = data['Sub-Specialty']
        get_sub_specialty = list(b.unique())

        c = data['Location']
        get_location = list(c.unique())

        d = data['Role']
        get_role_data = list(d.unique())

        e = data['PGY']
        get_pgy = list(e.unique())

                             

       
        if request.POST.getlist("Staff") and request.POST.getlist("Sub-Specialty") and request.POST.getlist("location") and request.POST.getlist("role") and request.POST.getlist("pgy"):
            staff = request.POST.getlist("Staff")
            sub_specialty = request.POST.getlist("sub_specialty")
            location = request.POST.getlist("location")
            role = request.POST.getlist("role")
            pgy = request.POST.getlist("pgy")
            data = data[data['Staff'].isin(staff)]
            data = data[data['Sub-Specialty'].isin(sub_specialty)]
            data = data[data['Location'].isin(location)]
            data = data[data['Role'].isin(role)]
            data = data[data['PGY'].isin(pgy)]
           
         

        if request.POST.getlist("Staff") and request.POST.getlist("Role") and not request.POST.getlist("PGY") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("location"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_role = request.POST.getlist("Role")
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Role"].isin(get_data_role)
            get_g_role = data[filter1 & filter2]
            get_g_staff =data[filter1 & filter2]
        
        if request.POST.getlist("Sub-Specialty") and request.POST.getlist("Location") and not request.POST.getlist("Role") and not request.POST.getlist("Staff") and not request.POST.getlist("PGY"):
            sub_specialty = request.POST.getlist("Sub-Specialty")
            get_data_location = request.POST.getlist("Location")
            filter1 = data["Sub-Specialty"].isin(sub_specialty)
            filter2 = data["Location"].isin(get_data_location)
            get_g_specialty = data[filter1 & filter2]
            get_g_location =data[filter1 & filter2]

        if request.POST.getlist("Staff") and request.POST.getlist("PGY") and not request.POST.getlist("Role") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("location"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_pgy = request.POST.getlist("PGY")
            get_data_pgy = list(map(int, get_data_pgy))
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["PGY"].isin(get_data_pgy)
            get_g_role = data[filter1 & filter2]
            get_g_pgy =data[filter1 & filter2]
        
        if request.POST.getlist("Location") and request.POST.getlist("PGY") and not request.POST.getlist("Role") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("Staff"):
            get_data_location = request.POST.getlist("Location")
            get_data_pgy = request.POST.getlist("PGY")
            get_data_pgy = list(map(int, get_data_pgy))
            filter1 = data["Location"].isin(get_data_location)
            filter2 = data["PGY"].isin(get_data_pgy)
            get_g_location = data[filter1 & filter2]
            get_g_pgy =data[filter1 & filter2]

        
        if request.POST.getlist("Sub-Specialty") and request.POST.getlist("PGY") and not request.POST.getlist("Role") and not request.POST.getlist("Staff") and not request.POST.getlist("Location"):
            sub_specialty = request.POST.getlist("Sub-Specialty")
            get_data_pgy= request.POST.getlist("PGY")

            get_data_pgy = list(map(int, get_data_pgy))
            filter1 = data["Sub-Specialty"].isin(sub_specialty)
            filter2 = data["PGY"].isin(get_data_pgy)
            get_g_specialty = data[filter1 & filter2]
            get_g_pgy =data[filter1 & filter2]



        if request.POST.getlist("Staff") and request.POST.getlist("Location") and not request.POST.getlist("Role") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("PGY"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_location = request.POST.getlist("Location")
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Location"].isin(get_data_location)
            get_g_staff = data[filter1 & filter2]
            get_g_location =data[filter1 & filter2]

        
        if request.POST.getlist("Staff") and request.POST.getlist("Sub-Specialty") and not request.POST.getlist("Location") and not request.POST.getlist("Role") and not request.POST.getlist("PGY"):
            get_data_staff = request.POST.getlist("Staff")
            sub_specialty = request.POST.getlist("Sub-Specialty")
            filter1 = data["Staff"].isin(get_data_staff)
   
 
            filter4 = data['Sub-Specialty'].isin(sub_specialty)
            get_g_staff = data[filter1 & filter4]
            get_g_specialty =data[filter1 & filter4]

        if request.POST.getlist("Staff") and request.POST.getlist("Role") and request.POST.getlist("Sub-Specialty"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_role = request.POST.getlist("Role")
            get_data_specialty = request.POST.getlist("Sub-Specialty")
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Role"].isin(get_data_role)
            filter4= data["Sub-Specialty"].isin(get_data_specialty)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_specialty =data[filter1 & filter2 & filter4]
        

        if request.POST.getlist("Staff") and request.POST.getlist("PGY") and request.POST.getlist("Sub-Specialty"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_pgy = request.POST.getlist("PGY")
            get_data_pgy = list(map(int, get_data_pgy))
            get_data_specialty = request.POST.getlist("Sub-Specialty")

            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["PGY"].isin(get_data_pgy)
            filter4= data["Sub-Specialty"].isin(get_data_specialty)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_pgy =data[filter1 & filter2 & filter4]
        
        if request.POST.getlist("Staff") and request.POST.getlist("Location") and request.POST.getlist("PGY"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_location = request.POST.getlist("Location")
            get_data_pgy = request.POST.getlist("PGY")
            get_data_pgy = list(map(int, get_data_pgy))

            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Location"].isin(get_data_location)
            filter4= data["PGY"].isin(get_data_pgy)
            
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_location = data[filter1 & filter2 & filter4]
            get_g_pgy =data[filter1 & filter2 & filter4]

        if request.POST.getlist("Staff") and request.POST.getlist("Role") and request.POST.getlist("Location"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_role = request.POST.getlist("Role")
            get_data_location = request.POST.getlist("Location")
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Role"].isin(get_data_role)
            filter4= data["Location"].isin(get_data_location)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_location =data[filter1 & filter2 & filter4]
        
        if request.POST.getlist("Staff") and request.POST.getlist("Sub-Specialty") and request.POST.getlist("Location"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_specialty = request.POST.getlist("Sub-Specialty")
            get_data_location = request.POST.getlist("Location")
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Sub-Specialty"].isin(get_data_specialty)
            filter4= data["Location"].isin(get_data_location)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_specialty =data[filter1 & filter2 & filter4]

        if request.POST.getlist("Sub-Specialty") and request.POST.getlist("Role")  and not request.POST.getlist("Staff"):

            get_data_specialty = request.POST.getlist("Sub-Specialty")
            get_data_role = request.POST.getlist("Role")
            filter4= data["Sub-Specialty"].isin(get_data_specialty)
            filter2 = data["Role"].isin(get_data_role)
            get_g_role = data[filter2 & filter4]
            get_g_specialty =data[filter2 & filter4]

        if request.POST.getlist("Location") and request.POST.getlist("Role") and not request.POST.getlist("Staff") and not request.POST.getlist("Sub-Specialty"):

            get_data_role = request.POST.getlist("Role")
            get_data_location = request.POST.getlist("Location")
            filter3 = data["Location"].isin(get_data_location)
            filter2 = data["Role"].isin(get_data_role)
            get_g_role = data[filter2 & filter3]
            get_g_location =data[filter2 & filter3]

        if request.POST.getlist("Location") and request.POST.getlist("Role") and request.POST.getlist("Staff"):
            get_data_location = request.POST.getlist("Location")
            get_data_role = request.POST.getlist("Role")
            get_data_staff = request.POST.getlist("Staff")
            filter1 = data["Location"].isin(get_data_location)
            filter2 = data["Role"].isin(get_data_role)
            filter4= data["Staff"].isin(get_data_staff)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_location =data[filter1 & filter2 & filter4]
            get_g_staff =data[filter1 & filter2 & filter4]

        if request.POST.getlist("Location") and request.POST.getlist("Role") and request.POST.getlist("Sub-Specialty"):
            get_data_location = request.POST.getlist("Location")
            get_data_role = request.POST.getlist("Role")
            get_data_specialty = request.POST.getlist("Sub-Specialty")
            filter1 = data["Location"].isin(get_data_location)
            filter2 = data["Role"].isin(get_data_role)
            filter4= data["Sub-Specialty"].isin(get_data_specialty)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_location =data[filter1 & filter2 & filter4]
            get_g_specialty =data[filter1 & filter2 & filter4]

        if request.POST.getlist("Location") and request.POST.getlist("Role") and request.POST.getlist("PGY"):
            get_data_location = request.POST.getlist("Location")
            get_data_role = request.POST.getlist("Role")
            get_data_pgy = request.POST.getlist("PGY")
            filter1 = data["Location"].isin(get_data_location)
            filter2 = data["Role"].isin(get_data_role)
            get_data_pgy = list(map(int, get_data_pgy))
            filter4= data["PGY"].isin(get_data_pgy)
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_location =data[filter1 & filter2 & filter4]
            get_g_pgy =data[filter1 & filter2 & filter4]

        if request.POST.getlist("Staff") and request.POST.getlist("Role") and request.POST.getlist("PGY"):
            get_data_staff = request.POST.getlist("Staff")
            get_data_role = request.POST.getlist("Role")
            get_data_pgy = request.POST.getlist("PGY")
            get_data_pgy = list(map(int, get_data_pgy))
            filter1 = data["Staff"].isin(get_data_staff)
            filter2 = data["Role"].isin(get_data_role)
            filter4= data["PGY"].isin(get_data_pgy)
            
            get_g_staff =data[filter1 & filter2 & filter4]
            get_g_role = data[filter1 & filter2 & filter4]
            get_g_pgy =data[filter1 & filter2 & filter4]

        if request.POST.getlist("PGY") and request.POST.getlist("Role") and not request.POST.getlist("Location") and not request.POST.getlist("Staff"):
    
            get_data_p = request.POST.getlist("PGY")
            get_data_role = request.POST.getlist("Role")
            get_data_pgy = [i.replace(".0", "") for i in get_data_p]
        
            for i in range(0, len(get_data_pgy)):
                get_data_pgy[i] = int(get_data_pgy[i])
            filter5 = data["PGY"].isin(get_data_pgy)

            filter2 = data["Role"].isin(get_data_role)
            get_g_role = data[filter2 & filter5]
            get_g_pgy =data[filter2 & filter5]

        if request.POST.getlist("PGY") and request.POST.getlist("Role") and request.POST.getlist("Staff") and request.POST.getlist("Location") and request.POST.getlist("Sub-Specialty"):
            get_data_staff = request.POST.getlist("Staff")
            filter1 = data["Staff"].isin(get_data_staff)

            get_data_role = request.POST.getlist("Role")
            filter2 = data["Role"].isin(get_data_role)

            get_data_location = request.POST.getlist("Location")
            filter3 = data["Location"].isin(get_data_location)

            get_data_specialty = request.POST.getlist("Sub-Specialty")
            filter4= data["Sub-Specialty"].isin(get_data_specialty)

            get_data_pgy = [i.replace(".0", "") for i in request.POST.getlist("PGY")]
        
            for i in range(0, len(get_data_pgy)):
                get_data_pgy[i] = int(get_data_pgy[i])
            filter5 = data["PGY"].isin(get_data_pgy)

            get_g_pgy =data[filter1 & filter2 & filter3 & filter4 & filter5]
            get_g_role = data[filter1 & filter2 & filter3 & filter4 & filter5]
            get_g_staff =data[filter1 & filter2 & filter3 & filter4 & filter5]
            get_g_location =data[filter1 & filter2 & filter3 & filter4 & filter5]
            get_g_specialty =data[filter1 & filter2 & filter3 & filter4 & filter5]



        if request.POST.getlist("Role") and request.POST.getlist("Staff") and request.POST.getlist("Location") and request.POST.getlist("Sub-Specialty"):
            get_data_staff = request.POST.getlist("Staff")
            filter1 = data["Staff"].isin(get_data_staff)

            get_data_role = request.POST.getlist("Role")
            filter2 = data["Role"].isin(get_data_role)

            get_data_location = request.POST.getlist("Location")
            filter3 = data["Location"].isin(get_data_location)

            get_data_specialty = request.POST.getlist("Sub-Specialty")
            filter4= data["Sub-Specialty"].isin(get_data_specialty)

           

            get_g_role = data[filter1 & filter2 & filter3 & filter4 ]
            get_g_staff =data[filter1 & filter2 & filter3 & filter4 ]
            get_g_location =data[filter1 & filter2 & filter3 & filter4 ]
            get_g_specialty =data[filter1 & filter2 & filter3 & filter4 ]


        

        if request.POST.getlist("Staff") and not request.POST.getlist("Role") and not request.POST.getlist("Location") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("PGY"):
         
            get_data_staff = request.POST.getlist("Staff")
            filter1 = data["Staff"].isin(get_data_staff)
            get_g_staff = data[filter1]

        if request.POST.getlist("Role") and not request.POST.getlist("Staff") and not request.POST.getlist("Location") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("PGY"):
     
            get_data_role = request.POST.getlist("Role")
            filter2 = data["Role"].isin(get_data_role)
            get_g_role = data[filter2]

        if request.POST.getlist("Location") and not request.POST.getlist("Role") and not request.POST.getlist("Staff") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("PGY"):
            
            get_data_location = request.POST.getlist("Location")
            filter3 = data["Location"].isin(get_data_location)
            get_g_location = data[filter3]

        if request.POST.getlist("Sub-Specialty") and not request.POST.getlist("Role") and not request.POST.getlist("Location") and not request.POST.getlist("Staff") and not request.POST.getlist("PGY"):
            get_data_specialty = request.POST.getlist("Sub-Specialty")
            filter4= data["Sub-Specialty"].isin(get_data_specialty)
            get_g_specialty = data[filter4]

        
           
        if request.POST.getlist("PGY") and not request.POST.getlist("Role") and not request.POST.getlist("Location") and not request.POST.getlist("Sub-Specialty") and not request.POST.getlist("Staff"):
            get_data_p = request.POST.getlist("PGY")
            get_data_pgy = [i.replace(".0", "") for i in get_data_p]
        
            for i in range(0, len(get_data_pgy)):
                get_data_pgy[i] = int(get_data_pgy[i])
            get_g_pgy = data[data["PGY"].isin(get_data_pgy)]
            print(get_g_pgy)

        context_di = {}

        staff = {}

        for i in data['Staff']:

            context_di[i] = data[data['Staff'] == i]

            staff[i] = dict(Counter(context_di[i]['Role']))


        if set(['Role','Sub-Specialty', 'Location','Date']).issubset(data.columns):
    
            save_obj = Graphs(user=request.user, upload=upload_file)
            save_obj.save()
           

            only_final_data = []

            for d in staff:
                final_data = {
                    "name": d,
                    "Primary Surgeon": staff[d].get('Primary Surgeon', 0),
                    "First Assist": staff[d].get('First Assist', 0),
                    "Secondary Assist": staff[d].get('Secondary Assist', 0),
                }
            
                only_final_data.append(final_data)

                

            datasets = [
                {
                "label": 'Primary Surgeon',
                "backgroundColor": '#398AB9',
                "data": []
                },
                {
                "label": 'First Assist',
                "backgroundColor": '#D8D2CB',
                "data": []
                },
                {
                "label": 'Secondary Assist',
                "backgroundColor": '#1A374D',
                "data": []
                }
            ]

            lists = []
            #update the data
            for i in only_final_data:
                total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                i['Total'] = total
                lists.append(i)
               
            name_set = []
            shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)
            for i in range(len(shorted_data)):
                name_set.append(shorted_data[i]['name'])


            p_list = []
            f_list = []
            s_list = []
            for d in shorted_data:
                s_list.append(d['Secondary Assist'])
                f_list.append(d['First Assist'])
                p_list.append(d['Primary Surgeon'])
               

            for dataset in datasets:
                if dataset['label'] == 'Primary Surgeon':
                    dataset['data'] = p_list
                elif dataset['label'] == 'First Assist':
                    dataset['data'] = f_list
                elif dataset['label'] == 'Secondary Assist':
                    dataset['data'] = s_list
          
            role = []
            specialty_chart = []
            site =[]
        
            if request.POST.getlist("PGY"):

                #ADD ROLE

                role=[]
                for i in get_g_pgy['Role']:
                    role.append(i)

                
                dashboard11 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard11.items(), key=lambda x:x[1],reverse=True)))
                keys = list(dashboard1.keys())
                values = list(dashboard1.values())
               

                #ADD SPECIALTY



                specialty_chart = {}
                roles = {}
                for i in get_g_pgy['Role']:
                    specialty_chart[i] = get_g_pgy[get_g_pgy['Role'] == i]
                    roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

                get_data_label = []
            
                for key, value in roles.items():
                
                    dat = {
                 
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        
                    }
                    get_data_label.append(dat)

                get_data = []

                for key, value in roles.items():
                
                    dat = {
                        "label": key,
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                    }
                    get_data.append(dat)

                shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)


                labels_data = list(get_data_label[0].keys())


                grt_Data = []

                for i in shorted_data_list:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    grt_Data.append(d)    


                #ADD SITE

                site_chart = {}
                roles_list = {}
                for i in get_g_pgy['Role']:
                    site_chart[i] = get_g_pgy[get_g_pgy['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))


                get_data_label = []

                for key, value in roles_list.items():
                
                    dat = {
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data_label.append(dat)
                    
                get_data = []

                for key, value in roles_list.items():
                
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)




                # key1 = list(roles_list.values())
                try:
                    labels_site = list(get_data_label[0].keys())
                except:
                    labels_site = []
    


                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 



                #ADD DATE

                context_dict = {}
                roles = {}
                for i in get_g_pgy['Date'].dt.strftime('%b-%Y'):
                    context_dict[i] =  get_g_pgy[get_g_pgy['Date'].dt.strftime('%b-%Y') == i]
                    roles[i] = dict(Counter(context_dict[i]['Role']))
     
                _data = {}
               
                for r in roles:
                    _data_set = {
                        'Secondary Assist': roles[r].get('Secondary Assist', 0),
                        'First Assist': roles[r].get('First Assist', 0),
                        'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                    }
                    
                    _data[r] = _data_set
            

                labels = list(roles.keys())
                role_keys = []
                raw_data = []

                for role in _data:

                    role_keys = (list(_data[role].keys()))
                    raw_data.append(_data[role])

                final_data = []
                for key in role_keys:
                    data = []
                    for raw in raw_data:
                        data.append(raw[key])
                        
                    final_data.append({
                    "label": key,
                    "data": data,
                    "borderColor": color_line_chart(key),
                    "fill": "true",
                    })
            
                final_final_data = {
                    "labels": labels,
                    "datasets": final_data,
                    
                }

                coded_case = []
                for i in get_g_pgy['Coded Case']:
                    coded_case.append(i)
            
                dashboard63 = dict(Counter(coded_case))
                dashboard23 = (dict(sorted(dashboard63.items(), key=lambda x:x[1],reverse=True)))
                #b add code of staff filter

                context_dib = {}

                staffs = {}

                for i in get_g_pgy['Staff']:
                    context_dib[i] =  get_g_pgy[get_g_pgy['Staff'] == i]
                    staffs[i] = dict(Counter(context_dib[i]['Role']))

                
   
                only_final_data = []

                for d in staffs:
                    final_data = {
                        "name": d,
                        "Primary Surgeon": staffs[d].get('Primary Surgeon', 0),
                        "First Assist": staffs[d].get('First Assist', 0),
                        "Secondary Assist": staffs[d].get('Secondary Assist', 0),
                    }
                
                    only_final_data.append(final_data)

                datasets = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]

                lists = []
            #update the data
                for i in only_final_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    lists.append(i)
                
                name_set = []
                shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)
                for i in range(len(shorted_data)):
                    name_set.append(shorted_data[i]['name'])

                p_list = []
                f_list = []
                s_list = []
                for d in shorted_data:

                    s_list.append(d['Secondary Assist'])
                    f_list.append(d['First Assist'])
                    p_list.append(d['Primary Surgeon'])

                for dataset in datasets:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_list

                        # end of code

                total_cases = len(get_g_pgy)
                current_year = date.today().year
                get_date = get_g_pgy[get_g_pgy['Date'].dt.year == current_year] 

                date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
                count_of_current_year = 0
                for u in list(date_gt):
                    if u == True:
                        count_of_current_year+=1
                    else:
                        pass
                tdate_equal =  get_g_pgy[get_g_pgy['Date'] == str(current_year)+'-'+'07-01'] 
                this_year_ps = date_equal['Role'] == 'Primary Surgeon' 

                count_of_current = 0
                for i in this_year_ps:
                    if i == True:
                        count_of_current+=1
                    else:
                        pass
                today = date.today()
                first = today.replace(day=1)
                lastMonth = first - timedelta(days=1)
                last_months = lastMonth.strftime("%Y-%m")
                this_year_month = get_g_pgy[get_g_pgy['Date'].dt.to_period('M') == last_months]
                get_g = this_year_month['Role'] == 'Primary Surgeon'
                count_of_last_month = 0
                for i in get_g:
                    if i == True:
                        count_of_last_month+=1
                    else:
                        pass


                today = date.today()
                datem = today.strftime("%Y-%m")
            
                this_year_this_month = get_g_pgy[get_g_pgy['Date'].dt.to_period('M') == datem]
                get_role = this_year_this_month['Role'] == 'Primary Surgeon'
                count_of_this_month = 0
                for i in get_role:
                    if i == True:
                        count_of_this_month+=1
                    else:
                        pass

                                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                speciality_label = []
                for i in get_g_pgy['Sub-Specialty']:
                    context_dib[i] =  get_g_pgy[get_g_pgy['Sub-Specialty'] == i]
                    if i not in speciality_label:
                        speciality_label.append(i)
                    sub_specialty[i] = dict(Counter(context_dib[i]['Role']))

                for sub in sub_specialty:
                    final_sub_specialty = {
                            "name": sub,
                            "Primary Surgeon": sub_specialty[sub].get('Primary Surgeon', 0),
                            "First Assist": sub_specialty[sub].get('First Assist', 0),
                            "Secondary Assist": sub_specialty[sub].get('Secondary Assist', 0),
                        }
                    
                    sub_specialty_raw.append(final_sub_specialty)
                
                datas = [
                        {
                        "label": 'Primary Surgeon',
                        "backgroundColor": '#398AB9',
                        "data": []
                        },
                        {
                        "label": 'First Assist',
                        "backgroundColor": '#D8D2CB',
                        "data": []
                        },
                        {
                        "label": 'Secondary Assist',
                        "backgroundColor": '#1A374D',
                        "data": []
                        }
                    ]
                
                for i in sub_specialty_raw:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)

                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list

            # Driver code by kapil END
                            #Driver Code for location start
                            ## Driver code by kapil START
                datas_loc = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                polish_location_data = []
                location_data = {}
                location_raw = []
                location_label = []
                for i in get_g_pgy['Location']:
                    context_dib[i] =  get_g_pgy[get_g_pgy['Location'] == i]
                    if i not in location_label:
                        location_label.append(i)
                    location_data[i] = dict(Counter(context_dib[i]['Role']))

                for loc in location_data:
                    final_location_specialty = {
                            "name": sub,
                            "Primary Surgeon": location_data[loc].get('Primary Surgeon', 0),
                            "First Assist": location_data[loc].get('First Assist', 0),
                            "Secondary Assist": location_data[loc].get('Secondary Assist', 0),
                        }
                    
                    location_raw.append(final_location_specialty)

                                
                for tot in location_raw:
                    total = tot['Primary Surgeon'] + tot['First Assist'] + tot['Secondary Assist']
                    tot['Total'] = total
                    polish_location_data.append(tot)

                shorted_loc_data = sorted(polish_location_data, key=lambda x: x['Total'], reverse=True)
                p_loc_list = []
                f_loc_list = []
                s_loc_list = []
                for d in shorted_loc_data:
                    p_loc_list.append(d['Primary Surgeon']) 
                    f_loc_list.append(d['First Assist']) 
                    s_loc_list.append(d['Secondary Assist']) 
                for dataset in datas_loc:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_loc_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_loc_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_loc_list

            ## Driver code by kapil END
            #Driver Code for location End
                try:
                    if request.user.is_superuser:
                        users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                    else:
                        sups = Supervisor.objects.filter(username=request.user.username)[0]
                        users = NewUser.objects.filter(supervisor=sups.id)
                except:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)

                context = {"keys":keys,"values":values,"labels_data":speciality_label, "final_data_specialty_chart":datas,"labels_site":location_label,"granted_Data":datas_loc,
                 'final_data': final_final_data, 'labels': labels,"final_data_list":datasets,"names":name_set, "get_staff":get_staff,"get_role_data":get_role_data, "get_sub_specialty":get_sub_specialty,"get_pgy":get_pgy, "get_location":get_location, "dashboard23":dashboard23, "total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month,
                 "users":users}
                return render(request, 'home/sample.html', context)

           

            if request.POST.getlist("Role"):
                # add Role

                role =[]
                
                for i in get_g_role['Role']:
                    role.append(i)
                
                dashboard111 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard111.items(), key=lambda x:x[1],reverse=True)))

                keys = list(dashboard1.keys())
                values = list(dashboard1.values())

            # add specialty
            
                specialty_chart = {}
                roles = {}
                for i in get_g_role['Role']:
                    specialty_chart[i] = get_g_role[get_g_role['Role'] == i]
                    roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

                get_data_label = []
            
                for key, value in roles.items():
                
                    dat = {
                 
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        
                    }
                    get_data_label.append(dat)

                get_data = []

                for key, value in roles.items():
                
                    dat = {
                        "label": key,
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                    }
                    get_data.append(dat)

                shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)

                try:
                    labels_data = list(get_data_label[0].keys())
                except:
                    labels_data = []


                grt_Data = []

                for i in shorted_data_list:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    grt_Data.append(d)   


            # add location

                # site=[]
                # for i in get_g_role['Location']:
                #     site.append(i)
                
                # dashboard61 = dict(Counter(site))
                # dashboard6 = (dict(sorted(dashboard61.items(), key=lambda x:x[1],reverse=True)))

                # keys2 = list(dashboard6.keys())
                # values2 = list(dashboard6.values())

                site_chart = {}
                roles_list = {}
                for i in get_g_role['Role']:
                    site_chart[i] = get_g_role[get_g_role['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))


                get_data_label = []

                for key, value in roles_list.items():
                
                    dat = {
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data_label.append(dat)
                    
                get_data = []

                for key, value in roles_list.items():
                
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)
                try:
                    labels_site = list(get_data_label[0].keys())
                except:
                    labels_site = []
                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 

            # add staff
                
                context_dib = {}

                staffs = {}

                for i in get_g_role['Staff']:
                    context_dib[i] =  get_g_role[get_g_role['Staff'] == i]
                    staffs[i] = dict(Counter(context_dib[i]['Role']))


                only_final_data = []

                for d in staffs:
                    final_data = {
                        "name": d,
                        "Primary Surgeon": staffs[d].get('Primary Surgeon', 0),
                        "First Assist": staffs[d].get('First Assist', 0),
                        "Secondary Assist": staffs[d].get('Secondary Assist', 0),
                    }
                
                    only_final_data.append(final_data)

                   

                datasets = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                lists = []
                for i in only_final_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    lists.append(i)
                    
                name_set = []

                shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)

                for i in range(len(shorted_data)):

                    name_set.append(shorted_data[i]['name'])

                p_list = []
                f_list = []
                s_list = []
                for d in shorted_data:
                
                    s_list.append(d['Secondary Assist'])
                    f_list.append(d['First Assist'])
                    p_list.append(d['Primary Surgeon'])

                for dataset in datasets:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_list
                        #end code of staff filter

                # add date
                
                context_dict = {}
                roles = {}
                for i in get_g_role['Date'].dt.strftime('%b-%Y'):
                    context_dict[i] =  get_g_role[get_g_role['Date'].dt.strftime('%b-%Y') == i]
                    roles[i] = dict(Counter(context_dict[i]['Role']))
                
                _data = {}
               
                for r in roles:
                    _data_set = {
                        'Secondary Assist': roles[r].get('Secondary Assist', 0),
                        'First Assist': roles[r].get('First Assist', 0),
                        'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                    }
                    
                    _data[r] = _data_set
            

                labels = list(roles.keys())
                role_keys = []
                raw_data = []

                for role in _data:

                    role_keys = (list(_data[role].keys()))
                    raw_data.append(_data[role])

                final_data = []
                for key in role_keys:
                    data = []
                    for raw in raw_data:
                        data.append(raw[key])
                        
                    final_data.append({
                    "label": key,
                    "data": data,
                    "borderColor": color_line_chart(key),
                    "fill": "true",
                    })

                final_final_data = {
                    "labels": labels,
                    "datasets": final_data,
                    
                }

                coded_case = []
                for i in get_g_role['Coded Case']:
                    coded_case.append(i)
            
                dashboard23 = dict(Counter(coded_case))

                total_cases = len(get_g_role)
                current_year = date.today().year
                get_date = get_g_role[get_g_role['Date'].dt.year == current_year] 
                date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
                count_of_current_year = 0
                for u in list(date_gt):
                    if u == True:
                        count_of_current_year+=1
                    else:
                        pass
                date_equal =  get_g_role[get_g_role['Date'] == str(current_year)+'-'+'07-01'] 
                this_year_ps = date_equal['Role'] == 'Primary Surgeon' 

                count_of_current = 0
                for i in this_year_ps:
                    if i == True:
                        count_of_current+=1
                    else:
                        pass
                today = date.today()
                first = today.replace(day=1)
                lastMonth = first - timedelta(days=1)
                last_months = lastMonth.strftime("%Y-%m")
                this_year_month = get_g_role[get_g_role['Date'].dt.to_period('M') == last_months]
                get_g = this_year_month['Role'] == 'Primary Surgeon'
                count_of_last_month = 0
                for i in get_g:
                    if i == True:
                        count_of_last_month+=1
                    else:
                        pass


                today = date.today()
                datem = today.strftime("%Y-%m")
        
                this_year_this_month = get_g_role[get_g_role['Date'].dt.to_period('M') == datem]
                get_role = this_year_this_month['Role'] == 'Primary Surgeon'
                count_of_this_month = 0
                for i in get_role:
                    if i == True:
                        count_of_this_month+=1
                    else:
                        pass
                                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                speciality_label = []
                for i in get_g_role['Sub-Specialty']:
                    context_dib[i] =  get_g_role[get_g_role['Sub-Specialty'] == i]
                    if i not in speciality_label:
                        speciality_label.append(i)
                    sub_specialty[i] = dict(Counter(context_dib[i]['Role']))

                for sub in sub_specialty:
                    final_sub_specialty = {
                            "name": sub,
                            "Primary Surgeon": sub_specialty[sub].get('Primary Surgeon', 0),
                            "First Assist": sub_specialty[sub].get('First Assist', 0),
                            "Secondary Assist": sub_specialty[sub].get('Secondary Assist', 0),
                        }
                    
                    sub_specialty_raw.append(final_sub_specialty)
                
                datas = [
                        {
                        "label": 'Primary Surgeon',
                        "backgroundColor": '#398AB9',
                        "data": []
                        },
                        {
                        "label": 'First Assist',
                        "backgroundColor": '#D8D2CB',
                        "data": []
                        },
                        {
                        "label": 'Secondary Assist',
                        "backgroundColor": '#1A374D',
                        "data": []
                        }
                    ]
                
                for i in sub_specialty_raw:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)

                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list

            # Driver code by kapil END
                            #Driver Code for location start
                            ## Driver code by kapil START
                datas_loc = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                polish_location_data = []
                location_data = {}
                location_raw = []
                location_label = []
                for i in get_g_role['Location']:
                    context_dib[i] =  get_g_role[get_g_role['Location'] == i]
                    if i not in location_label:
                        location_label.append(i)
                    location_data[i] = dict(Counter(context_dib[i]['Role']))

                for loc in location_data:
                    final_location_specialty = {
                            "name": sub,
                            "Primary Surgeon": location_data[loc].get('Primary Surgeon', 0),
                            "First Assist": location_data[loc].get('First Assist', 0),
                            "Secondary Assist": location_data[loc].get('Secondary Assist', 0),
                        }
                    
                    location_raw.append(final_location_specialty)

                                
                for tot in location_raw:
                    total = tot['Primary Surgeon'] + tot['First Assist'] + tot['Secondary Assist']
                    tot['Total'] = total
                    polish_location_data.append(tot)

                shorted_loc_data = sorted(polish_location_data, key=lambda x: x['Total'], reverse=True)
                p_loc_list = []
                f_loc_list = []
                s_loc_list = []
                for d in shorted_loc_data:
                    p_loc_list.append(d['Primary Surgeon']) 
                    f_loc_list.append(d['First Assist']) 
                    s_loc_list.append(d['Secondary Assist']) 
                for dataset in datas_loc:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_loc_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_loc_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_loc_list

            ## Driver code by kapil END
            #Driver Code for location End

                try:
                    if request.user.is_superuser:
                        users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                    else:
                        sups = Supervisor.objects.filter(username=request.user.username)[0]
                        users = NewUser.objects.filter(supervisor=sups.id)
                except:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)

                context = {"keys":keys,"values":values,"labels_data":speciality_label,"final_data_specialty_chart":datas,"labels_site":location_label,"granted_Data":datas_loc,
                    'final_data': final_final_data, 'labels': labels,"final_data_list":datasets,"names":name_set, "get_staff":get_staff,"get_role_data":get_role_data, "get_sub_specialty":get_sub_specialty,"get_pgy":get_pgy, "get_location":get_location, "dashboard23":dashboard23, "total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month,
                    "users":users}
                return render(request, 'home/sample.html', context)
                
            else:
                role =[]
                for i in data['Role']:
                    role.append(i)
                
                dashboard22 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard22.items(), key=lambda x:x[1],reverse=True)))

                keys = list(dashboard1.keys())
                values = list(dashboard1.values())
            

                coded_case = []
                for i in data['Coded Case']:
                    coded_case.append(i)
            
                dashboard31 = dict(Counter(coded_case))
                dashboard23 = (dict(sorted(dashboard31.items(), key=lambda x:x[1],reverse=True)))


            
            if request.POST.getlist("Sub-Specialty"):
             
                specialty_chart = {}
                roles = {}
                for i in get_g_specialty['Role']:
                    specialty_chart[i] = get_g_specialty[get_g_specialty['Role'] == i]
                    roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))


                get_data_label = []
            
                for key, value in roles.items():
                    
            
                    dat = {
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                      
                    }
                    get_data_label.append(dat)
                

                get_data = []
            
                for key, value in roles.items():
                    
            
                    dat = {
                        "label": key,
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                    }
                    get_data.append(dat)

                shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)
            
                
                # key1 = list(roles.values())
                try:
                    labels_data = list(get_data_label[0].keys())
                except:
                    labels_data = []
                
            
            
                grt_Data = []

                for i in shorted_data_list:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    grt_Data.append(d)    



                site_chart = {}
                roles_list = {}
                for i in get_g_specialty['Role']:
                    site_chart[i] = get_g_specialty[get_g_specialty['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))


                get_data_label = []

                for key, value in roles_list.items():
                
                    dat = {
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data_label.append(dat)
                    
                get_data = []

                for key, value in roles_list.items():
                
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)




                # key1 = list(roles_list.values())
                try:
                    labels_site = list(get_data_label[0].keys())
                except:
                    labels_site = []
    


                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 

                # for i in get_g_specialty['Location']:
                #     site.append(i)
                
                # dashboard16 = dict(Counter(site))
                # dashboard6 = (dict(sorted(dashboard16.items(), key=lambda x:x[1],reverse=True)))

                # keys2 = list(dashboard6.keys())
                # values2 = list(dashboard6.values())


                #get staff filter
                context_dib = {}

                staffs = {}

                for i in get_g_specialty['Staff']:
                    context_dib[i] =  get_g_specialty[get_g_specialty['Staff'] == i]
                    staffs[i] = dict(Counter(context_dib[i]['Role']))

              
             
                only_final_data = []

                for d in staffs:
                    final_data = {
                        "name": d,
                        "Primary Surgeon": staffs[d].get('Primary Surgeon', 0),
                        "First Assist": staffs[d].get('First Assist', 0),
                        "Secondary Assist": staffs[d].get('Secondary Assist', 0),
                    }
                
                    only_final_data.append(final_data)

                

                datasets = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]

                lists = []
                for i in only_final_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    lists.append(i)
                    
                name_set = []

                shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)

                for i in range(len(shorted_data)):
                    
                    name_set.append(shorted_data[i]['name'])

                p_list = []
                f_list = []
                s_list = []
                for d in shorted_data:
        
                    s_list.append(d['Secondary Assist'])
                    f_list.append(d['First Assist'])
                    p_list.append(d['Primary Surgeon'])

                for dataset in datasets:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_list
                        # end of get staff filter


                role=[]
                for i in get_g_specialty['Role']:
                    role.append(i)
                
                dashboard121 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard121.items(), key=lambda x:x[1],reverse=True)))
                keys = list(dashboard1.keys())
                values = list(dashboard1.values())
                

                
                context_dict = {}
                roles = {}
                for i in get_g_specialty['Date'].dt.strftime('%b-%Y'):
                    context_dict[i] =  get_g_specialty[get_g_specialty['Date'].dt.strftime('%b-%Y') == i]
                    roles[i] = dict(Counter(context_dict[i]['Role']))

    
                _data = {}
               
                for r in roles:
                    _data_set = {
                        'Secondary Assist': roles[r].get('Secondary Assist', 0),
                        'First Assist': roles[r].get('First Assist', 0),
                        'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                    }
                    
                    _data[r] = _data_set
            

                labels = list(roles.keys())
                role_keys = []
                raw_data = []

                for role in _data:

                    role_keys = (list(_data[role].keys()))
                    raw_data.append(_data[role])

                final_data = []
                for key in role_keys:
                    data = []
                    for raw in raw_data:
                        data.append(raw[key])
                        
                    final_data.append({
                    "label": key,
                    "data": data,
                    "borderColor": color_line_chart(key),
                    "fill": "true",
                    })
            
                final_final_data = {
                    "labels": labels,
                    "datasets": final_data,
                }

                coded_case = []
                for i in get_g_specialty['Coded Case']:
                    coded_case.append(i)

                total_cases = len(get_g_specialty)
                current_year = date.today().year
                get_date = get_g_specialty[get_g_specialty['Date'].dt.year == current_year] 

                date_gt = get_date['Date'].dt.to_period('M') > '2022-06'
                count_of_current_year = 0
                for u in list(date_gt):
                    if u == True:
                        count_of_current_year+=1
                    else:
                        pass
                date_equal =  get_g_specialty[get_g_specialty['Date'] == str(current_year)+'-'+'07-01'] 
                this_year_ps = date_equal['Role'] == 'Primary Surgeon'

                count_of_current = 0
                for i in this_year_ps:
                    if i == True:
                        count_of_current+=1
                    else:
                        pass
                today = date.today()
                first = today.replace(day=1)
                lastMonth = first - timedelta(days=1)
                last_months = lastMonth.strftime("%Y-%m")
                this_year_month = get_g_specialty[get_g_specialty['Date'].dt.to_period('M') == last_months]
                get_g = this_year_month['Role'] == 'Primary Surgeon'
                count_of_last_month = 0
                for i in get_g:
                    if i == True:
                        count_of_last_month+=1
                    else:
                        pass


                today = date.today()
                datem = today.strftime("%Y-%m")
              
                this_year_this_month = get_g_specialty[get_g_specialty['Date'].dt.to_period('M') == datem]
                get_role = this_year_this_month['Role'] == 'Primary Surgeon'
                count_of_this_month = 0
                for i in get_role:
                    if i == True:
                        count_of_this_month+=1
                    else:
                        pass


                dashboard223 = dict(Counter(coded_case))
                dashboard23 = (dict(sorted(dashboard223.items(), key=lambda x:x[1],reverse=True)))


                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                speciality_label = []
                for i in get_g_specialty['Sub-Specialty']:
                    context_dib[i] =  get_g_specialty[get_g_specialty['Sub-Specialty'] == i]
                    if i not in speciality_label:
                        speciality_label.append(i)
                    sub_specialty[i] = dict(Counter(context_dib[i]['Role']))

                for sub in sub_specialty:
                    final_sub_specialty = {
                            "name": sub,
                            "Primary Surgeon": sub_specialty[sub].get('Primary Surgeon', 0),
                            "First Assist": sub_specialty[sub].get('First Assist', 0),
                            "Secondary Assist": sub_specialty[sub].get('Secondary Assist', 0),
                        }
                    
                    sub_specialty_raw.append(final_sub_specialty)
                
                datas = [
                        {
                        "label": 'Primary Surgeon',
                        "backgroundColor": '#398AB9',
                        "data": []
                        },
                        {
                        "label": 'First Assist',
                        "backgroundColor": '#D8D2CB',
                        "data": []
                        },
                        {
                        "label": 'Secondary Assist',
                        "backgroundColor": '#1A374D',
                        "data": []
                        }
                    ]
                
                for i in sub_specialty_raw:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)

                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list

            # Driver code by kapil END
                


                #Driver code for procedures of sub-speciality start
                # labels_site = []
                # for site_filter in sub_specialty.keys():
                #     labels_site.append(site_filter)

                            # Driver code by kapil END
                
            #Driver Code for location start
                            ## Driver code by kapil START
                datas_loc = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                polish_location_data = []
                location_data = {}
                location_raw = []
                location_label = []
                for i in get_g_specialty['Location']:
                    context_dib[i] =  get_g_specialty[get_g_specialty['Location'] == i]
                    if i not in location_label:
                        location_label.append(i)
                    location_data[i] = dict(Counter(context_dib[i]['Role']))

                for loc in location_data:
                    final_location_specialty = {
                            "name": sub,
                            "Primary Surgeon": location_data[loc].get('Primary Surgeon', 0),
                            "First Assist": location_data[loc].get('First Assist', 0),
                            "Secondary Assist": location_data[loc].get('Secondary Assist', 0),
                        }
                    
                    location_raw.append(final_location_specialty)

                                
                for tot in location_raw:
                    total = tot['Primary Surgeon'] + tot['First Assist'] + tot['Secondary Assist']
                    tot['Total'] = total
                    polish_location_data.append(tot)

                shorted_loc_data = sorted(polish_location_data, key=lambda x: x['Total'], reverse=True)
                p_loc_list = []
                f_loc_list = []
                s_loc_list = []
                for d in shorted_loc_data:
                    p_loc_list.append(d['Primary Surgeon']) 
                    f_loc_list.append(d['First Assist']) 
                    s_loc_list.append(d['Secondary Assist']) 
                for dataset in datas_loc:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_loc_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_loc_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_loc_list

            ## Driver code by kapil END
            #Driver Code for location End
                try:
                    if request.user.is_superuser:
                        users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                    else:
                        sups = Supervisor.objects.filter(username=request.user.username)[0]
                        users = NewUser.objects.filter(supervisor=sups.id)
                except:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                context = {"keys":keys,"values":values, "labels_data":speciality_label, "final_data_specialty_chart":datas,"granted_Data":datas_loc, "labels_site":location_label, 'final_data': final_final_data, 'labels': labels,"final_data_list":datasets,"names":name_set, "get_staff":get_staff,"get_role_data":get_role_data,"get_pgy":get_pgy, "get_sub_specialty":get_sub_specialty, "get_location":get_location, "dashboard23":dashboard23,"total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month,'procedure_site_speciality':labels_site,"users":users}
                return render(request, 'home/sample.html', context)
                

                
            else:
                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                roles={}
                total_roles = []
                total_sub=[]
                raw_list=[]
                for i in data['Role']:
                    sub_specialty[i] = data[data['Role'] == i]
                    roles[i] = dict(Counter(sub_specialty[i]['Sub-Specialty']))

                for single_role in roles.keys():
                    total_roles.append(single_role)

                for sub in total_roles:
                    for m in roles[sub].keys():
                        if m not in total_sub:
                            total_sub.append(m)

                i = 0
                while i<len(total_sub):
                    final_data = {
                                    "name": total_sub[i],
                                    # "Primary Surgeon": roles['Primary Surgeon'].get(total_sub[i],0),
                                    "Primary Surgeon": roles.get('Primary Surgeon', {}).get(total_sub[i],0),
                                    "First Assist": roles.get('First Assist', {}).get(total_sub[i],0),
                                    # "First Assist": roles['First Assist'].get(total_sub[i],0),
                                    "Secondary Assist": roles['Secondary Assist'].get(total_sub[i],0)
                                }
                    raw_list.append(final_data)
                    i=i+1
                
                for i in raw_list:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)
                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 

                datas = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list

                #Driver code for location

                raw_site_data = []
                polish_site_data = []
                roles_site = []
                total_site = []
                roles_loc = {}
                loc_site = {}

                for i in data['Role']:
                    loc_site[i] = data[data['Role'] == i]
                    roles_loc[i] = dict(Counter(loc_site[i]['Location']))

                for role_location in roles_loc.keys():
                    roles_site.append(role_location)
    
                for sub in roles_site:
                    for m in roles_loc[sub].keys():
                        if m not in total_site:
                            total_site.append(m)

                i = 0  
                while i<len(total_site):
                    final_site_data = {
                    "name": total_site[i],
                    "Primary Surgeon": roles_loc.get('Primary Surgeon', {}).get(total_site[i],0),
                    "First Assist": roles_loc.get('First Assist', {}).get(total_site[i],0),
                    "Secondary Assist": roles_loc.get('Secondary Assist', {}).get(total_site[i],0)
                    }
                    raw_site_data.append(final_site_data)
                    i= i+1

                for i in raw_site_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polish_site_data.append(i)

                shorted_site_data = sorted(polish_site_data, key=lambda x: x['Total'], reverse=True)
                p_site_list = []
                f_site_list = []
                s_site_list = []
                for d in shorted_site_data:
                    p_site_list.append(d['Primary Surgeon']) 
                    f_site_list.append(d['First Assist']) 
                    s_site_list.append(d['Secondary Assist']) 

                datas_site = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                for dataset in datas_site:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_site_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_site_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_site_list

                

            # Driver code by kapil END

        #     # Driver code by kapil END
                #ADD CODE HERE
                # specialty_chart = {}
                # roles = {}
                # for i in data['Role']:
                #     specialty_chart[i] = data[data['Role'] == i]
                #     roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

                # get_data = []
          
                # for key, value in roles.items():
        
                #     dat = {
                #         "label": key,
                #         "Trauma": value.get("Trauma",0),
                #         "Arthroplasty": value.get("Arthroplasty",0),
                #         "Pediatrics": value.get("Pediatrics",0),
                #         "Sports": value.get("Sports",0),
                #         "Foot and Ankle": value.get("Foot and Ankle",0),
                #         "Upper Extremity": value.get("Upper Extremity",0),
                #         "General": value.get("General",0),
                #         "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                #     }
                #     get_data.append(dat)

                # shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)
          
            
                # key1 = list(roles.values())
                # labels_data = list(key1[0].keys())
        
         
                # grt_Data = []

                # for i in shorted_data_list:
                #     d ={
                #             "label": i.get("label"),
                #             "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                #             "backgroundColor": color_line_chart(i.get("label"))
                #         }
                #     grt_Data.append(d)    
                # print(grt_Data,"grt_Data")
       
                

          
            if request.POST.getlist("Location"):

                role=[]
                for i in get_g_location['Role']:
                    role.append(i)
                
                dashboard121 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard121.items(), key=lambda x:x[1],reverse=True)))

                keys = list(dashboard1.keys())
                values = list(dashboard1.values())
                # add staff filter

            # Procedure by surgeon
                context_dib = {}

                staffs = {}

                for i in get_g_location['Staff']:
                    context_dib[i] =  get_g_location[get_g_location['Staff'] == i]
                    staffs[i] = dict(Counter(context_dib[i]['Role']))


            
                only_final_data = []

                for d in staffs:
                    final_data = {
                        "name": d,
                        "Primary Surgeon": staffs[d].get('Primary Surgeon', 0),
                        "First Assist": staffs[d].get('First Assist', 0),
                        "Secondary Assist": staffs[d].get('Secondary Assist', 0),
                    }
                
                    only_final_data.append(final_data)

                datasets = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]

                lists = []
                for i in only_final_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    lists.append(i)
                    
                name_set = []

                shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)

                for i in range(len(shorted_data)):
                    
                    name_set.append(shorted_data[i]['name'])


                p_list = []
                f_list = []
                s_list = []
                for d in shorted_data:
    
                    s_list.append(d['Secondary Assist'])
                    f_list.append(d['First Assist'])
                    p_list.append(d['Primary Surgeon'])

                for dataset in datasets:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_list


                # add specialty filter

                specialty_chart = {}
                roles = {}
                for i in get_g_location['Role']:
                    specialty_chart[i] = get_g_location[get_g_location['Role'] == i]
                    roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

                get_data_label = []
            
                for key, value in roles.items():
                
                    dat = {
                 
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        
                    }
                    get_data_label.append(dat)

                get_data = []

                for key, value in roles.items():
                
                    dat = {
                        "label": key,
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                    }
                    get_data.append(dat)

                shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)

                try:
                    labels_data = list(get_data_label[0].keys())
                except:
                    labels_data = []


                grt_Data = []

                for i in shorted_data_list:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    grt_Data.append(d)    
           
           # add location filter

                site_chart = {}
                roles_list = {}
                for i in get_g_location['Role']:
                    site_chart[i] = get_g_location[get_g_location['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))


                get_data_label = []

                for key, value in roles_list.items():
                
                    dat = {
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data_label.append(dat)
                    
                get_data = []

                for key, value in roles_list.items():
                
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)

                # key1 = list(roles_list.values())
                try:
                    labels_site = list(get_data_label[0].keys())
                except:
                    labels_site = []
  
                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 
            # date filter


                context_dict = {}
                roles = {}
                for i in data['Date'].dt.strftime('%b-%Y'):
                    context_dict[i] =  get_g_location[get_g_location['Date'].dt.strftime('%b-%Y') == i]
                    roles[i] = dict(Counter(context_dict[i]['Role']))
                
                
                _data = {}
               
                for r in roles:
                    _data_set = {
                        'Secondary Assist': roles[r].get('Secondary Assist', 0),
                        'First Assist': roles[r].get('First Assist', 0),
                        'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                    }
                    
                    _data[r] = _data_set
            

                labels = list(roles.keys())
                role_keys = []
                raw_data = []

                for role in _data:

                    role_keys = (list(_data[role].keys()))
                    raw_data.append(_data[role])

                final_data = []
                for key in role_keys:
                    data = []
                    for raw in raw_data:
                        data.append(raw[key])
                        
                    final_data.append({
                    "label": key,
                    "data": data,
                    "borderColor": color_line_chart(key),
                    "fill": "true",
                    })


                final_final_data = {
                    "labels": labels,
                    "datasets": final_data,
                }

                coded_case = []
                for i in get_g_location['Coded Case']:
                    coded_case.append(i)
            
                dashboard43 = dict(Counter(coded_case))
                dashboard23 = (dict(sorted(dashboard43.items(), key=lambda x:x[1],reverse=True)))

                total_cases = len(get_g_location)
                current_year = date.today().year
                get_date = get_g_location[get_g_location['Date'].dt.year == current_year] 

                date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
                count_of_current_year = 0
                for u in list(date_gt):
                    if u == True:
                        count_of_current_year+=1
                    else:
                        pass
                date_equal =  get_g_location[get_g_location['Date'] == str(current_year)+'-'+'07-01'] 
                this_year_ps = date_equal['Role'] == 'Primary Surgeon' 

                count_of_current = 0
                for i in this_year_ps:
                    if i == True:
                        count_of_current+=1
                    else:
                        pass
                today = date.today()
                first = today.replace(day=1)
                lastMonth = first - timedelta(days=1)
                last_months = lastMonth.strftime("%Y-%m")
                this_year_month = get_g_location[get_g_location['Date'].dt.to_period('M') == last_months]
                get_g = this_year_month['Role'] == 'Primary Surgeon'
                count_of_last_month = 0
                for i in get_g:
                    if i == True:
                        count_of_last_month+=1
                    else:
                        pass


                today = date.today()
                datem = today.strftime("%Y-%m")
          
                this_year_this_month = get_g_location[get_g_location['Date'].dt.to_period('M') == datem]
                get_role = this_year_this_month['Role'] == 'Primary Surgeon'
                count_of_this_month = 0
                for i in get_role:
                    if i == True:
                        count_of_this_month+=1
                    else:
                        pass
                                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                speciality_label = []
                for i in get_g_location['Sub-Specialty']:
                    context_dib[i] =  get_g_location[get_g_location['Sub-Specialty'] == i]
                    if i not in speciality_label:
                        speciality_label.append(i)
                    sub_specialty[i] = dict(Counter(context_dib[i]['Role']))

                for sub in sub_specialty:
                    final_sub_specialty = {
                            "name": sub,
                            "Primary Surgeon": sub_specialty[sub].get('Primary Surgeon', 0),
                            "First Assist": sub_specialty[sub].get('First Assist', 0),
                            "Secondary Assist": sub_specialty[sub].get('Secondary Assist', 0),
                        }
                    
                    sub_specialty_raw.append(final_sub_specialty)
                
                datas = [
                        {
                        "label": 'Primary Surgeon',
                        "backgroundColor": '#398AB9',
                        "data": []
                        },
                        {
                        "label": 'First Assist',
                        "backgroundColor": '#D8D2CB',
                        "data": []
                        },
                        {
                        "label": 'Secondary Assist',
                        "backgroundColor": '#1A374D',
                        "data": []
                        }
                    ]
                
                for i in sub_specialty_raw:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)

                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list


                                
            #Driver Code for location start
                            ## Driver code by kapil START
                datas_loc = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                polish_location_data = []
                location_data = {}
                location_raw = []
                location_label = []
                for i in get_g_location['Location']:
                    context_dib[i] =  get_g_location[get_g_location['Location'] == i]
                    if i not in location_label:
                        location_label.append(i)
                    location_data[i] = dict(Counter(context_dib[i]['Role']))

                for loc in location_data:
                    final_location_specialty = {
                            "name": sub,
                            "Primary Surgeon": location_data[loc].get('Primary Surgeon', 0),
                            "First Assist": location_data[loc].get('First Assist', 0),
                            "Secondary Assist": location_data[loc].get('Secondary Assist', 0),
                        }
                    
                    location_raw.append(final_location_specialty)

                                
                for tot in location_raw:
                    total = tot['Primary Surgeon'] + tot['First Assist'] + tot['Secondary Assist']
                    tot['Total'] = total
                    polish_location_data.append(tot)

                shorted_loc_data = sorted(polish_location_data, key=lambda x: x['Total'], reverse=True)
                p_loc_list = []
                f_loc_list = []
                s_loc_list = []
                for d in shorted_loc_data:
                    p_loc_list.append(d['Primary Surgeon']) 
                    f_loc_list.append(d['First Assist']) 
                    s_loc_list.append(d['Secondary Assist']) 
                for dataset in datas_loc:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_loc_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_loc_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_loc_list

            ## Driver code by kapil END
            #Driver Code for location End

            # Driver code by kapil END
                try:
                    if request.user.is_superuser:
                        users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                    else:
                        sups = Supervisor.objects.filter(username=request.user.username)[0]
                        users = NewUser.objects.filter(supervisor=sups.id)
                except:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)

                context = {"keys":keys,"values":values,"final_data_list":datasets,"names":name_set,"labels_data":speciality_label, "final_data_specialty_chart":datas,"labels_site":location_label,"granted_Data":datas_loc,
                 'final_data': final_final_data, 'labels': labels, "get_staff":get_staff,"get_pgy":get_pgy,"get_role_data":get_role_data, "get_sub_specialty":get_sub_specialty, "get_location":get_location, "dashboard23":dashboard23,"total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month,
                 "users":users
                }
                return render(request, 'home/sample.html', context)
                
            else:
                site_chart = {}
                roles_list = {}
                for i in data['Role']:
                    site_chart[i] = data[data['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))


                get_data = []

                for key, value in roles_list.items():
                
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)



                key1 = list(roles_list.values())
                labels_site = list(key1[0].keys())
    


                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 
                    
                

            if request.POST.getlist("Staff"):

                context_dib = {}

                staffs = {}

                for i in get_g_staff['Staff']:
                    context_dib[i] =  get_g_staff[get_g_staff['Staff'] == i]
                    staffs[i] = dict(Counter(context_dib[i]['Role']))


                only_final_data = []

                for d in staffs:
                    final_data = {
                        "name": d,
                        "Primary Surgeon": staffs[d].get('Primary Surgeon', 0),
                        "First Assist": staffs[d].get('First Assist', 0),
                        "Secondary Assist": staffs[d].get('Secondary Assist', 0),
                    }
                
                    only_final_data.append(final_data)

                datasets = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                lists = []
                for i in only_final_data:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    lists.append(i)
                    
                name_set = []

                shorted_data = sorted(lists, key=lambda x: x['Total'], reverse=True)

                for i in range(len(shorted_data)):
                    
                    name_set.append(shorted_data[i]['name'])

                p_list = []
                f_list = []
                s_list = []
                for d in shorted_data:
         
                    s_list.append(d['Secondary Assist'])
                    f_list.append(d['First Assist'])
                    p_list.append(d['Primary Surgeon'])
                
                


                for dataset in datasets:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_list
                    elif dataset['label'] == 'First Assist':
                        dataset['data'] = f_list
                    elif dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_list



                role=[]
                for i in get_g_staff['Role']:
                    role.append(i)
                
                dashboard212 = dict(Counter(role))
                dashboard1 = (dict(sorted(dashboard212.items(), key=lambda x:x[1],reverse=True)))

                keys = list(dashboard1.keys())
                values = list(dashboard1.values())
               
                # specialty_chart =[]
                # for i in get_g_staff['Sub-Specialty']:
                #     specialty_chart.append(i)
                
                # dashboard13 = dict(Counter(specialty_chart))
                # dashboard3 = (dict(sorted(dashboard13.items(), key=lambda x:x[1],reverse=True )))

                # keys1 = list(dashboard3.keys())
                # values1 = list(dashboard3.values())
                specialty_chart = {}
                roles = {}
                for i in get_g_staff['Role']:
                    specialty_chart[i] = get_g_staff[get_g_staff['Role'] == i]
                    roles[i] = dict(Counter(specialty_chart[i]['Sub-Specialty']))

                get_data_label = []
            
                for key, value in roles.items():
            
                    dat = {
                       
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        
                    }
                    get_data_label.append(dat)


                get_data = []
            
                for key, value in roles.items():
            
                    dat = {
                        "label": key,
                        "Trauma": value.get("Trauma",0),
                        "Arthroplasty": value.get("Arthroplasty",0),
                        "Pediatrics": value.get("Pediatrics",0),
                        "Sports": value.get("Sports",0),
                        "Foot and Ankle": value.get("Foot and Ankle",0),
                        "Upper Extremity": value.get("Upper Extremity",0),
                        "General": value.get("General",0),
                        "total": value.get("Trauma",0) + value.get("Arthroplasty",0) + value.get("Pediatrics",0) + value.get("Sports",0) + value.get("Foot and Ankle",0) + value.get("Upper Extremity",0) + value.get("General",0)
                    }
                    get_data.append(dat)

                shorted_data_list = sorted(get_data, key=lambda x: x['total'], reverse=True)
            
                
                try:
                    labels_data = list(get_data_label[0].keys())
                except:
                    labels_data = []
            
            
                grt_Data = []

                for i in shorted_data_list:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("Trauma"), i.get("Arthroplasty"), i.get("Pediatrics"), i.get("Sports"), i.get("Foot and Ankle"), i.get("Upper Extremity"), i.get("General")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    grt_Data.append(d)    
   

                site_chart = {}
                roles_list = {}
                for i in get_g_staff['Role']:
                    site_chart[i] = get_g_staff[get_g_staff['Role'] == i]
                    roles_list[i] = dict(Counter(site_chart[i]['Location']))
                
                get_data_label = []
            
                for key, value in roles_list.items():
            
                    dat = {

                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data_label.append(dat)
        


                get_data = []
            
                for key, value in roles_list.items():
            
                    dat = {
                        "label": key,
                        "HGH": value.get("HGH",0),
                        "JH": value.get("JH",0),
                        "MUMC": value.get("MUMC",0),
                        "SJH": value.get("SJH",0),
                    }
                    get_data.append(dat)
                
                try:
                    labels_site = list(get_data_label[0].keys())
                except:
                    labels_site = []
    
                
            
                granted_Data = []

                for i in get_data:
                    d ={
                            "label": i.get("label"),
                            "data": [i.get("HGH"), i.get("JH"), i.get("MUMC"), i.get("SJH")],
                            "backgroundColor": color_line_chart(i.get("label"))
                        }
                    granted_Data.append(d) 



                context_dict = {}
                roles = {}
                for i in data['Date'].dt.strftime('%b-%Y'):
                    context_dict[i] =  data[data['Date'].dt.strftime('%b-%Y') == i]
                    roles[i] = dict(Counter(context_dict[i]['Role']))
               
                _data = {}
               
                for r in roles:
                    _data_set = {
                        'Secondary Assist': roles[r].get('Secondary Assist', 0),
                        'First Assist': roles[r].get('First Assist', 0),
                        'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                    }
                    
                    _data[r] = _data_set
            

                labels = list(roles.keys())
                role_keys = []
                raw_data = []

                for role in _data:

                    role_keys = (list(_data[role].keys()))
                    raw_data.append(_data[role])

                final_data = []
                for key in role_keys:
                    data = []
                    for raw in raw_data:
                        data.append(raw[key])
                        
                    final_data.append({
                    "label": key,
                    "data": data,
                    "borderColor": color_line_chart(key),
                    "fill": "true",
                    })
                final_final_data = {
                    "labels": labels,
                    "datasets": final_data,
                    
                }

                coded_case = []
                for i in get_g_staff['Coded Case']:
                    coded_case.append(i)
            
                dashboard123 = dict(Counter(coded_case))
                dashboard23 = (dict(sorted(dashboard123.items(), key=lambda x:x[1],reverse=True)))

                total_cases = len(get_g_staff)
                current_year = date.today().year
                get_date = get_g_staff[get_g_staff['Date'].dt.year == current_year] 

                date_gt = get_date['Date'].dt.to_period('M') > str(current_year)+'-'+'06'
                count_of_current_year = 0
                for u in list(date_gt):
                    if u == True:
                        count_of_current_year+=1
                    else:
                        pass
                date_equal =  get_g_staff[get_g_staff['Date'] == str(current_year)+'-'+'07-01'] 
                this_year_ps = date_equal['Role'] == 'Primary Surgeon' 

                count_of_current = 0
                for i in this_year_ps:
                    if i == True:
                        count_of_current+=1
                    else:
                        pass
                today = date.today()
                first = today.replace(day=1)
                lastMonth = first - timedelta(days=1)
                last_months = lastMonth.strftime("%Y-%m")
                this_year_month = get_g_staff[get_g_staff['Date'].dt.to_period('M') == last_months]
                get_g = this_year_month['Role'] == 'Primary Surgeon'
                count_of_last_month = 0
                for i in get_g:
                    if i == True:
                        count_of_last_month+=1
                    else:
                        pass


                today = date.today()
                datem = today.strftime("%Y-%m")

                this_year_this_month = get_g_staff[get_g_staff['Date'].dt.to_period('M') == datem]
                get_role = this_year_this_month['Role'] == 'Primary Surgeon'
                count_of_this_month = 0
                for i in get_role:
                    if i == True:
                        count_of_this_month+=1
                    else:
                        pass
                                #Driver code for procedures of sub-speciality start
                                ## Driver code by kapil START
                sub_specialty = {}
                sub_specialty_raw = []
                polished_data = []
                speciality_label = []
                for i in get_g_staff['Sub-Specialty']:
                    context_dib[i] =  get_g_staff[get_g_staff['Sub-Specialty'] == i]
                    if i not in speciality_label:
                        speciality_label.append(i)
                    sub_specialty[i] = dict(Counter(context_dib[i]['Role']))

                for sub in sub_specialty:
                    final_sub_specialty = {
                            "name": sub,
                            "Primary Surgeon": sub_specialty[sub].get('Primary Surgeon', 0),
                            "First Assist": sub_specialty[sub].get('First Assist', 0),
                            "Secondary Assist": sub_specialty[sub].get('Secondary Assist', 0),
                        }
                    
                    sub_specialty_raw.append(final_sub_specialty)
                
                datas = [
                        {
                        "label": 'Primary Surgeon',
                        "backgroundColor": '#398AB9',
                        "data": []
                        },
                        {
                        "label": 'First Assist',
                        "backgroundColor": '#D8D2CB',
                        "data": []
                        },
                        {
                        "label": 'Secondary Assist',
                        "backgroundColor": '#1A374D',
                        "data": []
                        }
                    ]
                
                for i in sub_specialty_raw:
                    total = i['Primary Surgeon'] + i['First Assist'] + i['Secondary Assist']
                    i['Total'] = total
                    polished_data.append(i)

                shorted_data_list = sorted(polished_data, key=lambda x: x['Total'], reverse=True)
                p_spec_list = []
                f_spec_list = []
                s_spec_list = []
                for d in shorted_data_list:
                    p_spec_list.append(d['Primary Surgeon']) 
                    f_spec_list.append(d['First Assist']) 
                    s_spec_list.append(d['Secondary Assist']) 
                for dataset in datas:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_spec_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_spec_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_spec_list

            # Driver code by kapil END
                
            #Driver Code for location start
                            ## Driver code by kapil START
                datas_loc = [
                    {
                    "label": 'Primary Surgeon',
                    "backgroundColor": '#398AB9',
                    "data": []
                    },
                    {
                    "label": 'First Assist',
                    "backgroundColor": '#D8D2CB',
                    "data": []
                    },
                    {
                    "label": 'Secondary Assist',
                    "backgroundColor": '#1A374D',
                    "data": []
                    }
                ]
                polish_location_data = []
                location_data = {}
                location_raw = []
                location_label = []
                for i in get_g_staff['Location']:
                    context_dib[i] =  get_g_staff[get_g_staff['Location'] == i]
                    if i not in location_label:
                        location_label.append(i)
                    location_data[i] = dict(Counter(context_dib[i]['Role']))

                for loc in location_data:
                    final_location_specialty = {
                            "name": sub,
                            "Primary Surgeon": location_data[loc].get('Primary Surgeon', 0),
                            "First Assist": location_data[loc].get('First Assist', 0),
                            "Secondary Assist": location_data[loc].get('Secondary Assist', 0),
                        }
                    
                    location_raw.append(final_location_specialty)

                                
                for tot in location_raw:
                    total = tot['Primary Surgeon'] + tot['First Assist'] + tot['Secondary Assist']
                    tot['Total'] = total
                    polish_location_data.append(tot)

                shorted_loc_data = sorted(polish_location_data, key=lambda x: x['Total'], reverse=True)
                p_loc_list = []
                f_loc_list = []
                s_loc_list = []
                for d in shorted_loc_data:
                    p_loc_list.append(d['Primary Surgeon']) 
                    f_loc_list.append(d['First Assist']) 
                    s_loc_list.append(d['Secondary Assist']) 
                for dataset in datas_loc:
                    if dataset['label'] == 'Primary Surgeon':
                        dataset['data'] = p_loc_list
                    if dataset['label'] == 'First Assist':
                        dataset['data'] = f_loc_list
                    if dataset['label'] == 'Secondary Assist':
                        dataset['data'] = s_loc_list

            ## Driver code by kapil END
            #Driver Code for location End


                #Driver code for procedures of sub-speciality start
                try:
                    if request.user.is_superuser:
                        users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                    else:
                        sups = Supervisor.objects.filter(username=request.user.username)[0]
                        users = NewUser.objects.filter(supervisor=sups.id)
                except:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                context = {"keys":keys,"values":values, "labels_data":speciality_label, "final_data_specialty_chart":datas,"granted_Data":datas_loc,"labels_site":location_label,'final_data': final_final_data, 'labels': labels,"final_data_list":datasets,"names":name_set, "get_staff":get_staff,"get_role_data":get_role_data, "get_sub_specialty":get_sub_specialty,"get_pgy":get_pgy, "get_location":get_location, "dashboard23":dashboard23, "total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month
                ,"users":users}
                return render(request, 'home/sample.html', context)

            context_dict = {}
            roles = {}
            for i in data['Date'].dt.strftime('%b-%Y'):
                context_dict[i] =  data[data['Date'].dt.strftime('%b-%Y') == i]
                roles[i] = dict(Counter(context_dict[i]['Role']))
            
            _data = {}
               
            for r in roles:
                _data_set = {
                    'Secondary Assist': roles[r].get('Secondary Assist', 0),
                    'First Assist': roles[r].get('First Assist', 0),
                    'Primary Surgeon': roles[r].get('Primary Surgeon', 0)
                }
                
                _data[r] = _data_set
        
            labels = list(roles.keys())
            role_keys = []
            raw_data = []
            for role in _data:
                role_keys = (list(_data[role].keys()))
                raw_data.append(_data[role])
            final_data = []
            for key in role_keys:
                data = []
                for raw in raw_data:
                    data.append(raw[key])
                    
                final_data.append({
                "label": key,
                "data": data,
                "borderColor": color_line_chart(key),
                "fill": "true",
                })
        
            final_final_data = {
                   "labels": labels,
                "datasets": final_data,
             
            }
            try:
                if request.user.is_superuser:
                    users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
                else:
                    sups = Supervisor.objects.filter(username=request.user.username)[0]
                    users = NewUser.objects.filter(supervisor=sups.id)
            except:
                users = NewUser.objects.filter(is_superuser=False,is_supervisor=False)
            
            context = {"keys":keys,"values":values,"final_data_specialty_chart":datas, "labels_data":total_sub, 'granted_Data': datas_site, 'labels_site': total_site , 'final_data': final_final_data, 'labels': labels,"final_data_list":datasets,"names":name_set, "get_staff":get_staff,"get_role_data":get_role_data,"get_pgy":get_pgy, "get_sub_specialty":get_sub_specialty, "get_location":get_location, "dashboard23":dashboard23, "total_cases":total_cases, "count_of_current_year":count_of_current_year, "count_of_current":count_of_current,"count_of_last_month":count_of_last_month, 'count_of_this_month':count_of_this_month,'institute':institute, 'supervisor':supervisor,

            "users":users}
            if request.FILES.get('document'):
                messages.success(request, "File uploaded successfully")
                return render(request, 'home/sample.html', context)
            else:
                return render(request, 'home/sample.html', context)
            

        else:
            messages.error(request, "Invalid header in file")

            return render(request, 'home/sample.html')
    else:
        return render(request, 'home/sample.html')


def profile_pic(request):

    img = open('apps/home/anime3.png', 'rb')
    response = FileResponse(img)
    return response

    
def AddSupervisor(request):
    get_values= Supervisor.objects.all()
    institute = Institution.objects.all()
    print(institute,"------------")
   
    if request.method == "POST":
        username=request.POST.get("username")
        password = request.POST.get("password")
        get_institute_name =request.POST.get("get_institute_name")
        print(get_institute_name,"{{{}}}}}")
        confirm_password = request.POST.get("confirm_password")
        try:
            institue_name=Institution.objects.get(id=get_institute_name)
            print(institue_name,"--0-9-9-9")
        except Exception as e:
            print(e)
      
        if NewUser.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return redirect('/')
        else:
            username=username
        if password == confirm_password:
            user = NewUser.objects.create(username=username)
            user.password=make_password(password)
            user.is_supervisor=True
            user.save()
            supervisor = Supervisor.objects.create(username=username,institute=institue_name)
            messages.success(request, "Supervisor created successfully")
            return redirect("/supervisor")
        else:
            messages.error(request, "Password doesn't matched")
            return redirect("/supervisor")
    return render(request, 'home/add-supervisor.html', {"get_values":get_values,"institute":institute})

def AddInstitute(request):
    get_institute= Institution.objects.all()
    if request.method == "POST":
        institute=request.POST.get("institute")
        user = Institution.objects.create(institute=institute)
        messages.success(request, "Institute created successfully")
        return redirect("/institute")

    return render(request, 'home/add-institute.html', {"get_institute":get_institute})

def AddUser(request):
    get_user= NewUser.objects.filter(is_superuser=False, is_supervisor=False)
    institute = Institution.objects.all()
    supervisor = Supervisor.objects.all()
    if request.method =="POST":
        email=request.POST.get("email")
        username=request.POST.get("username")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
        institute=request.POST.get("institute")
        print(institute,"++++++++++")
        supervisor=request.POST.get("supervisor")
        password=request.POST.get("password")
        confirm_password=request.POST.get("confirm_password")
        try:
            obj_institute = Institution.objects.filter(institute=institute)
        except:
            messages.error(request, "Institute doesn't exists")
            return redirect('user')
        try:
            obj_supervisor = Supervisor.objects.get(username=supervisor)
        except:
            messages.error(request, "Supervisor doesn't exists")
            return redirect('user')

        if NewUser.objects.filter(username=username).exists():
            messages.error(request, "Username already exists")
            return redirect('user')
        else:
            username=username

        if NewUser.objects.filter(email=email).exists():
            messages.error(request, "Email already exists")
            return redirect('user')
        else:
            email=email
        if password == confirm_password:
            users = NewUser.objects.create(email=email,username=username,first_name=first_name,last_name=last_name,institute=obj_institute,supervisor=obj_supervisor)
            users.password=make_password(password)
            users.is_active=True
            users.save()
            messages.success(request, "User created successfully")
            return redirect("/user")
        else:
            messages.error(request, "Password doesn't matched")
            return redirect("/user")
    return render(request, 'home/add-user.html', {"get_user":get_user, "institute":institute, "supervisor":supervisor})

def UserProfile(request):

    if request.method =="POST":
 
        get_profile = NewUser.objects.get(username=request.user)
        get_profile.profile_photo=request.FILES.get('profile_photo')
        get_profile.save()
        messages.success(request, "Profile Updated Sucessfully")
        return redirect("/Profile")
    return render(request, 'home/user-profile.html')

def Procedure(request):
    try:
        new_user = NewUser.objects.get(username=request.user)
        get_file_data = Graphs.objects.filter(user=new_user).last()
        file_path = get_file_data.upload.path

        skipcols = ["PGY", "★","Notes","For Next Time","Case"]
        data=pd.read_excel(file_path,usecols=lambda x: x not in skipcols, sheet_name='RAW', header=2, parse_dates=True)
        pandas_data_frame= data.sort_values(by='Date')
        data = pd.DataFrame(pandas_data_frame)
        data.index = data.index + 1

        columns = list(data.columns.values)
        greek = data.to_html(classes="data",justify="left",index=1,table_id = "procedure_table")
    
    except:
        greek = ''
        data = ''


    return render(request, 'home/procedure.html', {"greek":greek,"data":data}) 


def delete(request,id):
    obj = NewUser.objects.get(id=id)
    obj.delete()
    return redirect('user')

def deletesupervisor(request,id):
    delsupervisor=Supervisor.objects.get(id=id)
    delsupervisor.delete()
    return redirect('user')

def deleteinstitute(request,id):
    delinstitute=Institution.objects.get(id=id)
    delinstitute.delete()
    return redirect('institute')